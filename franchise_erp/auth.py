import frappe
from frappe import _


def check_session_limit(login_manager):

    user = frappe.form_dict.get("usr")

    if not user or user.lower() in ["administrator", "guest"]:
        return

    active_sessions = frappe.db.count("Sessions", {"user": user})

    if active_sessions > 0:

        html_msg = _(
        """
        This user id is already logged in.<br>
        Click
        <button type="button" class="btn btn-primary"
        onclick="
        var user=document.getElementById('login_email').value;
        var pass=document.getElementById('login_password').value;

        frappe.call({
            method:'franchise_erp.auth.force_logout_and_login',
            args:{user:user},
            callback:function(){
                fetch('/api/method/login',{
                    method:'POST',
                    headers:{'Content-Type':'application/json'},
                    body:JSON.stringify({usr:user,pwd:pass})
                }).then(r=>r.json()).then(()=>{
                    window.location='/app';
                });
            }
        });
        ">
        Login
        </button>
        to login again.
        <br>
        Please note, by clicking this your existing session will be automatically logged out without saving.
        """
        )

        frappe.throw(html_msg, frappe.AuthenticationError)


@frappe.whitelist(allow_guest=True)
def force_logout_and_login(user):

    # ERPNext built-in session clear
    frappe.sessions.clear_sessions(user)

    frappe.db.commit()

    return {"status": "success"}





import frappe
from openpyxl import load_workbook


@frappe.whitelist()
def upload_serial_excel(file_url, supplier):

    serial_list = read_excel_serials(file_url)
    po_docs = get_purchase_orders(supplier)

    final_items = []
    errors = []
    gate_entries = set()

    for serial in serial_list:
        result = demo_serial(serial, po_docs, gate_entries)

        if result.get("item"):
            final_items.append(result["item"])

        if result.get("error"):
            errors.append(result["error"])

    # SAVE ONCE (IMPORTANT)
    for po in po_docs.values():
        po.save(ignore_permissions=True)

    return {
        "items": final_items,
        "errors": errors,
        "gate_entry_list": list(gate_entries)
    }


# -----------------------------
# READ EXCEL
# -----------------------------
def read_excel_serials(file_url):

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_path = file_doc.get_full_path()

    wb = load_workbook(file_path)
    sheet = wb.active

    serials = []

    for row in sheet.iter_rows(min_row=1, max_col=1):
        if row[0].value:
            serials.append(str(row[0].value).strip())

    return list(dict.fromkeys(serials))


# -----------------------------
# GET PO
# -----------------------------
def get_purchase_orders(supplier):

    po_list = frappe.get_all(
        "Purchase Order",
        filters={"supplier": supplier, "docstatus": 1},
        pluck="name"
    )

    return {po: frappe.get_doc("Purchase Order", po) for po in po_list}


# -----------------------------
# PROCESS SERIAL
# -----------------------------
def demo_serial(serial, po_docs, gate_entries):

    for po in po_docs.values():

        for item in po.items:

            pending_qty = item.qty - item.received_qty
            gate_no = get_gate_entry(item, gate_entries)

            if item.custom_unused_serials:
                res = handle_unused(serial, item, po.name, pending_qty, gate_no)
                if res:
                    return res

            if item.custom_generated_serials:
                res = handle_generated(serial, item, po.name, pending_qty, gate_no)
                if res:
                    return res

    return {"error": f"{serial} not found"}


# -----------------------------
# GATE ENTRY
# -----------------------------
def get_gate_entry(item, gate_entries):

    if not item.custom_incoming_logistic:
        return None

    il_doc = frappe.db.get_value(
        "Incoming Logistics",
        item.custom_incoming_logistic,
        ["gate_entry_no"],
        as_dict=1
    )

    if il_doc and il_doc.get("gate_entry_no"):
        gate_entries.add(il_doc["gate_entry_no"])
        return il_doc["gate_entry_no"]

    return None


# -----------------------------
# UNUSED
# -----------------------------
def handle_unused(serial, item, po_name, pending_qty, gate_no):

    unused = split_lines(item.custom_unused_serials)

    if serial not in unused:
        return None

    if pending_qty <= 0:
        return {"error": f"{serial} qty exceeded for {item.item_code}"}

    unused.remove(serial)

    used = split_lines(item.custom_used_serials)
    used.append(serial)

    item.custom_unused_serials = "\n".join(unused)
    item.custom_used_serials = "\n".join(used)

    return {"item": build_item(item, po_name, serial, gate_no)}


# -----------------------------
# GENERATED
# -----------------------------
def handle_generated(serial, item, po_name, pending_qty, gate_no):

    gen = split_lines(item.custom_generated_serials)

    if serial not in gen:
        return None

    if pending_qty <= 0:
        return {"error": f"{serial} qty exceeded for {item.item_code}"}

    return {"item": build_item(item, po_name, serial, gate_no)}


# -----------------------------
# BUILD ITEM
# -----------------------------
def build_item(item, po_name, serial, gate_no):

    return {
        "item_code": item.item_code,
        "item_name": item.item_name,
        "description": item.description,
        "qty": 1,
        "received_qty": 1,
        "uom": item.uom,
        "stock_uom": item.uom,
        "conversion_factor": 1,
        "rate": item.rate,
        "base_rate": item.rate,
        "purchase_order": po_name,
        "purchase_order_item": item.name,
        "serial_no": serial,
        "warehouse": item.warehouse,
        "custom_bulk_gate_entry": gate_no,
        "use_serial_batch_fields": 1
    }


def split_lines(text):
    if not text:
        return []
    return [x.strip() for x in text.split("\n") if x.strip()]
