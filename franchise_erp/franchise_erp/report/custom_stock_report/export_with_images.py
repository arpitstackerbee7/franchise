
import os

import frappe
from frappe.utils import flt, get_site_path
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from franchise_erp.franchise_erp.report.custom_stock_report.custom_stock_report import execute


@frappe.whitelist()
def export_custom_stock_report_with_images(filters=None, report_data=None):
    if isinstance(filters, str):
        filters = frappe.parse_json(filters)

    columns, fresh_data, _message = execute(filters)
    if report_data:
        data = frappe.parse_json(report_data) if isinstance(report_data, str) else report_data
    else:
        data = fresh_data
    # -----------------------------
    # Add Total Row
    # -----------------------------
    first_field = columns[0].get("fieldname")
    data = [
    row for row in data
    if not (isinstance(row, dict) and row.get(first_field) == "Total")
] 
    # -----------------------------
    # Add Total Row
    # -----------------------------
    if data:
      

        # First visible column
        total_row = {}
        total_row[first_field] = "Total"

        numeric_types = (
            "Currency",
            "Float",
            "Int",
            "Percent",
            "Decimal",
        )

        for col in columns:
            fieldname = col.get("fieldname")
            fieldtype = col.get("fieldtype")

            if (
                fieldname
                and fieldtype in numeric_types
            ):
                total_row[fieldname] = sum(
                    flt(row.get(fieldname))
                    for row in data
                    if isinstance(row, dict)
                )

        data.append(total_row)

    wb = Workbook()
    ws = wb.active
    ws.title = "Custom Stock Report"

    # -----------------------------
    # Headers
    # -----------------------------
    for col_idx, col in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col.get("label"))

    image_col_idx = None

    for idx, col in enumerate(columns):
        if col.get("fieldname") == "image":
            image_col_idx = idx
            break

    # -----------------------------
    # Data
    # -----------------------------
    for row_idx, row in enumerate(data, start=2):

        for col_idx, col in enumerate(columns, start=1):

            fieldname = col.get("fieldname")
            value = row.get(fieldname)

            if image_col_idx is not None and (col_idx - 1) == image_col_idx:

                # Don't try to insert image in Total row
                if (
                    value
                    and row.get(first_field) != "Total"
                ):
                    file_path = None

                    if value.startswith("/files/"):
                        file_path = get_site_path(
                            "public",
                            value.lstrip("/"),
                        )

                    elif value.startswith("/private/files/"):
                        file_path = get_site_path(
                            value.lstrip("/")
                        )

                    if file_path and os.path.exists(file_path):
                        try:
                            img = XLImage(file_path)
                            img.width = 60
                            img.height = 60
                            ws.add_image(
                                img,
                                f"{get_column_letter(col_idx)}{row_idx}",
                            )
                        except Exception:
                            pass

                ws.row_dimensions[row_idx].height = 50

            else:
                ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=value,
                )

    # -----------------------------
    # Column Widths
    # -----------------------------
    for col_idx, col in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)

        width = col.get("width") or 100

        ws.column_dimensions[col_letter].width = max(
            15,
            width / 7,
        )

    # -----------------------------
    # Save File
    # -----------------------------
    file_name = (
        f"custom_stock_report_"
        f"{frappe.utils.now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    temp_path = os.path.join(
        get_site_path("private", "files"),
        file_name,
    )

    wb.save(temp_path)

    with open(temp_path, "rb") as f:
        file_doc = frappe.get_doc(
            {
                "doctype": "File",
                "file_name": file_name,
                "is_private": 1,
                "content": f.read(),
            }
        )

        file_doc.save(ignore_permissions=True)

    os.remove(temp_path)

    return file_doc.file_url


